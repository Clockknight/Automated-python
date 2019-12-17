import os
import sys
import openpyxl
from openpyxl import workbook

#Take 3 argvs
if len(sys.argv) != 4 or str(sys.argv[3][-5:]) != '.xlsx':
    print("Input arguments were invalid. Please make sure that the input looks something like:\n\trowInserter.py 3 4 too_Compact.xlsx\n\tWhere the two integers are the rows you don't want moved, and the amount of rows you want the rest of the sheet moved down, respectively.")
baseLine = int(sys.argv[1])#first argv is what line is the baseLine
expansion = int(sys.argv[2])#second argv is the amount of rows that will be expanded
    #third argv is target xlsx file
    #Load book
book = openpyxl.load_workbook(str(sys.argv[3]))
sheet = book.active

#Get sheet dimensions
sheetWidth = sheet.max_column
sheetHeight = sheet.max_row
if baseLine >= sheetHeight:
    print("Your first integer was too great. There are only %d rows in %s." % (sheetHeight,str(sys.argv[3])))

#Copy values down the sheet
for i in range(1, sheetHeight-baseLine):
    for j in range(1, sheetWidth):
        sheet.cell(row=sheetHeight-i+1+expansion, column=j).value = sheet.cell(row=sheetHeight-i+1, column=j).value

#Clear cells for the rows needed
for i in range(1, expansion):
    for j in range(1, sheetWidth):
        sheet.cell(row=baseLine+i, column=j).value = ''

#Save
book.save('inserted'+str(sys.argv[3]))
